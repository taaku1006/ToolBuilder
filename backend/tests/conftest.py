"""Pytest fixtures for backend tests."""

from __future__ import annotations

import csv
import datetime
from collections.abc import Generator
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
import pytest_asyncio
from fastapi.testclient import TestClient
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from core.config import Settings
from db.engine import Base, get_db


# ---------------------------------------------------------------------------
# Settings / OpenAI mocks
# ---------------------------------------------------------------------------


@pytest.fixture
def mock_settings() -> Settings:
    """Return a Settings instance with test values."""
    return Settings(
        openai_api_key="test-api-key-12345",
        openai_model="gpt-4o",
        cors_origins="http://localhost:5173",
    )


@pytest.fixture
def mock_openai_response() -> dict:
    """Return a mock OpenAI JSON response payload."""
    return {
        "summary": "Excelファイルの集計処理",
        "python_code": (
            "import os\n"
            "import openpyxl\n\n"
            "INPUT_FILE = os.environ['INPUT_FILE']\n"
            "OUTPUT_DIR = os.environ['OUTPUT_DIR']\n\n"
            "# Excelファイルを読み込む\n"
            "wb = openpyxl.load_workbook(INPUT_FILE)\n"
            "ws = wb.active\n"
            "print('処理完了')\n"
        ),
        "steps": ["ファイルを読み込む", "データを集計する", "結果を保存する"],
        "tips": "INPUT_FILE環境変数にExcelファイルのパスを設定してください。",
    }


@pytest.fixture
def mock_openai_json_str(mock_openai_response: dict) -> str:
    """Return mock OpenAI response as JSON string."""
    import json

    return json.dumps(mock_openai_response, ensure_ascii=False)


@pytest.fixture
def mock_openai_client(mock_openai_json_str: str) -> Generator[MagicMock, None, None]:
    """Patch OpenAIClient.generate_code to return a mock JSON string."""
    with patch(
        "services.openai_client.OpenAI",
        autospec=True,
    ) as mock_openai_cls:
        mock_instance = MagicMock()
        mock_openai_cls.return_value = mock_instance

        mock_choice = MagicMock()
        mock_choice.message.content = mock_openai_json_str
        mock_instance.chat.completions.create.return_value = MagicMock(
            choices=[mock_choice]
        )
        yield mock_instance


@pytest.fixture
def test_client(
    mock_settings: Settings,
    mock_openai_client: MagicMock,
) -> Generator[TestClient, None, None]:
    """Return a FastAPI TestClient with mocked settings and OpenAI."""
    from core.deps import get_settings
    from main import app

    app.dependency_overrides[get_settings] = lambda: mock_settings

    with TestClient(app) as client:
        yield client

    app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# File fixtures — xlsx / csv
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Create a real xlsx file with known data for parser tests.

    Sheet name: Sales
    Columns: date (datetime), product (str), quantity (int), price (float), active (bool)
    Data rows: 5
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws.append(["date", "product", "quantity", "price", "active"])

    rows = [
        [datetime.date(2024, 1, 1), "Widget A", 10, 9.99, True],
        [datetime.date(2024, 1, 2), "Widget B", 5, 19.99, False],
        [datetime.date(2024, 1, 3), "Gadget X", 20, 4.50, True],
        [datetime.date(2024, 1, 4), "Gadget Y", 2, 99.00, True],
        [datetime.date(2024, 1, 5), "Widget A", 8, 9.99, False],
    ]
    for row in rows:
        ws.append(row)

    path = tmp_path / "sample.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def large_xlsx(tmp_path: Path) -> Path:
    """Create an xlsx with 50 data rows to test the 30-row preview cap."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BigSheet"

    ws.append(["id", "value"])
    for i in range(1, 51):
        ws.append([i, i * 10])

    path = tmp_path / "large.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    """Create an xlsx with two sheets: Sheet1 and Sheet2."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["a", "b"])
    ws1.append([1, 2])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["x", "y"])
    ws2.append([10, 20])

    path = tmp_path / "multi.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def empty_xlsx(tmp_path: Path) -> Path:
    """Create an xlsx with a sheet that has no data rows (only header or nothing)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"
    # No rows at all — completely empty sheet

    path = tmp_path / "empty.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def xlsx_with_nulls(tmp_path: Path) -> Path:
    """Create an xlsx where some cells are None/empty."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nulls"

    ws.append(["col_a", "col_b", "col_c"])
    ws.append([1, None, "hello"])
    ws.append([None, 2, None])
    ws.append([3, 4, "world"])

    path = tmp_path / "nulls.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_csv(tmp_path: Path) -> Path:
    """Create a real csv file with 3 data rows."""
    path = tmp_path / "sample.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name", "value"])
        writer.writerow([1, "Alice", 100])
        writer.writerow([2, "Bob", 200])
        writer.writerow([3, "Charlie", 300])
    return path


# ---------------------------------------------------------------------------
# Upload dir fixture
# ---------------------------------------------------------------------------


@pytest.fixture
def tmp_upload_dir(tmp_path: Path) -> Path:
    """Return a temporary directory used as the upload destination."""
    d = tmp_path / "uploads"
    d.mkdir()
    return d


# ---------------------------------------------------------------------------
# Upload test client
# ---------------------------------------------------------------------------


_TEST_DB_URL = "sqlite+aiosqlite:///:memory:"


@pytest_asyncio.fixture
async def _shared_test_engine():
    """In-memory SQLite engine shared across upload + skills tests."""
    engine = create_async_engine(_TEST_DB_URL, echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield engine
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
    await engine.dispose()


@pytest_asyncio.fixture
async def _shared_session_factory(_shared_test_engine):
    """Session factory for the shared in-memory engine."""
    return async_sessionmaker(
        bind=_shared_test_engine,
        class_=AsyncSession,
        expire_on_commit=False,
    )


@pytest.fixture
def upload_client(
    tmp_upload_dir: Path,
    mock_openai_client: MagicMock,
    _shared_session_factory,
) -> Generator[TestClient, None, None]:
    """Return a TestClient wired to use tmp_upload_dir, mocked OpenAI, and in-memory DB."""
    upload_settings = Settings(
        openai_api_key="test-api-key-12345",
        openai_model="gpt-4o",
        cors_origins="http://localhost:5173",
        upload_dir=str(tmp_upload_dir),
    )

    from core.deps import get_settings
    from main import app

    async def override_get_db():
        async with _shared_session_factory() as session:
            yield session

    app.dependency_overrides[get_settings] = lambda: upload_settings
    app.dependency_overrides[get_db] = override_get_db

    with TestClient(app) as client:
        yield client

    app.dependency_overrides.clear()
