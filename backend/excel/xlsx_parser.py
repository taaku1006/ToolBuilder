"""Excel / CSV file parser using openpyxl and pandas.

Returns immutable SheetInfo dataclasses — no mutation of parsed results.
"""

from __future__ import annotations

import csv
import datetime
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

_SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
_PREVIEW_ROWS = 30
_TYPE_INFERENCE_ROWS = 100

# Sheets with this many merged ranges are treated as output templates
_TEMPLATE_MERGE_THRESHOLD = 8
# Keywords that identify template sheets by name
_TEMPLATE_NAME_KEYWORDS = ("テンプレート", "template", "Template", "フォーム", "form")


@dataclass(frozen=True)
class SheetInfo:
    """Immutable summary of a single spreadsheet sheet."""

    name: str
    total_rows: int
    headers: list[str]
    types: dict[str, str]
    preview: list[dict[str, str | int | float | None]]
    # Merged cell ranges (e.g. "A1:J1", "A7:D7")
    merged_cells: tuple[str, ...] = ()
    # For template-like sheets: row-by-row structural map
    template_map: str = ""


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse_file(file_path: str) -> list[SheetInfo]:
    """Parse an xlsx / xls / csv file and return a SheetInfo for each sheet.

    Raises:
        FileNotFoundError: when the file does not exist.
        ValueError: when the file extension is not supported.
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = path.suffix.lower()
    if ext not in _SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file extension '{ext}'. "
            f"Allowed: {', '.join(sorted(_SUPPORTED_EXTENSIONS))}"
        )

    if ext == ".csv":
        return _parse_csv(path)
    return _parse_xlsx(path)


def build_file_context(
    sheets: list[SheetInfo],
    max_sample_rows: int = 3,
) -> str:
    """Build a human-readable text summary of file structure for prompt injection."""
    if not sheets:
        return ""

    parts: list[str] = []

    for sheet in sheets:
        lines: list[str] = [f"[Sheet: {sheet.name}]"]
        lines.append(f"Rows: {sheet.total_rows}")

        if sheet.merged_cells:
            lines.append(f"Merged cells: {', '.join(sheet.merged_cells)}")

        if sheet.template_map:
            lines.append(
                "⚠️ テンプレートシート検出: 結合セルあり。"
                "書き込み時は各結合範囲の左上セルのみに値を設定すること。"
            )
            lines.append("Template structure (行番号:セル範囲=値 | (空:データ書込み位置)):")
            lines.append(sheet.template_map)
        else:
            col_descriptions = ", ".join(
                f"{h} ({sheet.types.get(h, 'unknown')})" for h in sheet.headers
            )
            lines.append(f"Columns: {col_descriptions}")

            sample_count = min(max_sample_rows, len(sheet.preview))
            if sample_count > 0:
                lines.append(f"Sample rows ({sample_count}):")
                for row in sheet.preview[:sample_count]:
                    row_str = ", ".join(
                        f"{k}={v!r}" for k, v in row.items()
                    )
                    lines.append(f"  {row_str}")

        parts.append("\n".join(lines))

    return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _parse_xlsx(path: Path) -> list[SheetInfo]:
    """Parse all sheets of an xlsx/xls file using openpyxl + pandas."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    results: list[SheetInfo] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        merged: tuple[str, ...] = tuple(str(r) for r in ws.merged_cells.ranges)
        is_template = (
            any(kw in sheet_name for kw in _TEMPLATE_NAME_KEYWORDS)
            or len(merged) >= _TEMPLATE_MERGE_THRESHOLD
        )
        template_map = _build_template_map(ws) if is_template else ""

        if not rows:
            results.append(
                SheetInfo(
                    name=sheet_name,
                    total_rows=0,
                    headers=[],
                    types={},
                    preview=[],
                    merged_cells=merged,
                    template_map=template_map,
                )
            )
            continue

        header_row = rows[0]
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        data_rows = rows[1:]

        total_rows = len(data_rows)
        preview_rows = _build_preview(headers, data_rows[:_PREVIEW_ROWS])
        types = _infer_types_from_openpyxl(headers, data_rows[:_TYPE_INFERENCE_ROWS])

        results.append(
            SheetInfo(
                name=sheet_name,
                total_rows=total_rows,
                headers=headers,
                types=types,
                preview=preview_rows,
                merged_cells=merged,
                template_map=template_map,
            )
        )

    return results


def _build_template_map(ws) -> str:
    """Build a row-by-row structural map of a template/form sheet.

    Shows which cells contain labels and which are empty data-entry positions,
    taking merged cell ranges into account.
    """
    # Index merged ranges by their starting row
    merged_starting: dict[int, list] = defaultdict(list)
    # Track all (row, col) positions covered by any merge
    merge_coverage: set[tuple[int, int]] = set()

    for mr in ws.merged_cells.ranges:
        merged_starting[mr.min_row].append(mr)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_coverage.add((r, c))

    # Collect row numbers that have visible content or start a merged range
    relevant_rows: set[int] = set(merged_starting.keys())
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                relevant_rows.add(cell.row)

    lines: list[str] = []
    for row_num in sorted(relevant_rows):
        parts: list[str] = []
        col = 1
        max_col = ws.max_column or 1

        while col <= max_col:
            # Check if a merged range starts at this cell
            starts_here = [
                mr for mr in merged_starting.get(row_num, [])
                if mr.min_col == col
            ]
            if starts_here:
                mr = starts_here[0]
                col_start = get_column_letter(mr.min_col)
                col_end = get_column_letter(mr.max_col)
                range_str = f"{col_start}{mr.min_row}:{col_end}{mr.max_row}"
                val = ws.cell(mr.min_row, mr.min_col).value
                if val is not None:
                    parts.append(f"{range_str}={repr(val)}")
                else:
                    parts.append(f"{range_str}=(空:データ書込み位置)")
                col = mr.max_col + 1
            elif (row_num, col) in merge_coverage:
                # Covered by a merge that started elsewhere — skip
                col += 1
            else:
                cell = ws.cell(row_num, col)
                if cell.value is not None:
                    parts.append(f"{get_column_letter(col)}{row_num}={repr(cell.value)}")
                col += 1

        if parts:
            lines.append(f"  行{row_num}: " + " | ".join(parts))

    return "\n".join(lines)


def _parse_csv(path: Path) -> list[SheetInfo]:
    """Parse a CSV file using pandas, returning a single SheetInfo named 'Sheet1'."""
    df = pd.read_csv(str(path))
    headers = list(df.columns.astype(str))
    total_rows = len(df)

    # Build preview from raw rows
    preview_df = df.head(_PREVIEW_ROWS)
    preview_rows: list[dict[str, str | int | float | None]] = []
    for _, row in preview_df.iterrows():
        preview_rows.append(
            {h: _coerce_value(row[h]) for h in headers}
        )

    types = _infer_types_from_pandas(df.head(_TYPE_INFERENCE_ROWS))

    return [
        SheetInfo(
            name="Sheet1",
            total_rows=total_rows,
            headers=headers,
            types=types,
            preview=preview_rows,
        )
    ]


def _infer_types_from_openpyxl(
    headers: list[str],
    data_rows: list[tuple],
) -> dict[str, str]:
    """Infer column types from raw openpyxl cell values."""
    if not data_rows or not headers:
        return {h: "string" for h in headers}

    type_map: dict[str, str] = {}

    for col_idx, header in enumerate(headers):
        values = [
            row[col_idx] if col_idx < len(row) else None
            for row in data_rows
        ]
        non_null = [v for v in values if v is not None]

        if not non_null:
            type_map[header] = "string"
            continue

        type_map[header] = _detect_type_from_values(non_null)

    return type_map


def _detect_type_from_values(values: list) -> str:
    """Detect the dominant type from a list of non-null sample values."""
    date_types = (datetime.date, datetime.datetime)
    bool_count = sum(1 for v in values if isinstance(v, bool))
    date_count = sum(1 for v in values if isinstance(v, date_types) and not isinstance(v, bool))
    num_count = sum(1 for v in values if isinstance(v, (int, float)) and not isinstance(v, bool))

    total = len(values)
    threshold = total * 0.5

    if bool_count > threshold:
        return "boolean"
    if date_count > threshold:
        return "date"
    if num_count > threshold:
        return "number"
    return "string"


def _infer_types_from_pandas(df: pd.DataFrame) -> dict[str, str]:
    """Map pandas dtypes to simplified type strings."""
    type_map: dict[str, str] = {}
    for col in df.columns:
        dtype = df[col].dtype
        col_str = str(col)
        if pd.api.types.is_bool_dtype(dtype):
            type_map[col_str] = "boolean"
        elif pd.api.types.is_datetime64_any_dtype(dtype):
            type_map[col_str] = "date"
        elif pd.api.types.is_numeric_dtype(dtype):
            type_map[col_str] = "number"
        else:
            type_map[col_str] = "string"
    return type_map


def _build_preview(
    headers: list[str],
    data_rows: list[tuple],
) -> list[dict[str, str | int | float | None]]:
    """Convert raw openpyxl rows into a list of header-keyed dicts."""
    preview: list[dict[str, str | int | float | None]] = []
    for row in data_rows:
        row_dict: dict[str, str | int | float | None] = {}
        for col_idx, header in enumerate(headers):
            raw = row[col_idx] if col_idx < len(row) else None
            row_dict[header] = _coerce_value(raw)
        preview.append(row_dict)
    return preview


def _coerce_value(value: object) -> str | int | float | None:
    """Coerce a raw cell value to a JSON-serialisable type."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value  # type: ignore[return-value]
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, (datetime.date, datetime.datetime)):
        return str(value)
    return str(value)
