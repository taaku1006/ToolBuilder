"""Structured Excel comparator using openpyxl.

Produces a rich CompareReport JSON that:
- Scans for label cells by value (not by row position)
- Checks fill colors for conditional formatting
- Verifies sheet visibility (hidden/visible)
- Checks for extra output files

Designed to correctly evaluate cases where calculations are correct
but data is placed in the wrong rows.
"""

from __future__ import annotations

import glob
import json
import logging
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

_COLOR_FAMILIES = {
    "red":    lambda r, g, b: r > 180 and g < 150 and b < 150,
    "yellow": lambda r, g, b: r > 180 and g > 180 and b < 100,
    "green":  lambda r, g, b: r < 150 and g > 150 and b < 150,
    "blue":   lambda r, g, b: r < 150 and g < 150 and b > 180,
}


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------


@dataclass
class KeyCellResult:
    coord: str
    description: str
    expected: object
    actual: object
    match: bool


@dataclass
class ValueScanFieldResult:
    name: str
    expected: object
    actual: object
    match: bool


@dataclass
class ValueScanResult:
    label: str
    description: str
    found: bool
    found_at: str | None  # e.g. "A24"
    field_results: list[ValueScanFieldResult]


@dataclass
class ColorCheckResult:
    description: str
    row_range: tuple[int, int]
    col_range: tuple[int, int]
    expected_color_family: str
    cells_checked: int
    cells_matched: int
    match_rate: float
    passed: bool


@dataclass
class SheetStructureResult:
    missing_sheets: list[str]
    unexpected_sheets: list[str]
    row_counts: dict[str, dict]   # sheet -> {expected, actual}
    visibility: dict[str, dict]   # sheet -> {expected, actual}


@dataclass
class StructuredCompareReport:
    sheet_structure: SheetStructureResult
    key_cell_results: list[KeyCellResult]
    value_scan_results: list[ValueScanResult]
    color_check_results: list[ColorCheckResult]
    extra_file_results: list[dict]

    def to_dict(self) -> dict:
        return {
            "sheet_structure": asdict(self.sheet_structure),
            "key_cells": [asdict(r) for r in self.key_cell_results],
            "value_scan": [asdict(r) for r in self.value_scan_results],
            "color_checks": [asdict(r) for r in self.color_check_results],
            "extra_files": self.extra_file_results,
        }

    def to_json(self, indent: int = 2) -> str:
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=indent)

    def summary_text(self) -> str:
        """Build a human-readable summary for LLM consumption."""
        lines: list[str] = []

        ss = self.sheet_structure
        if ss.missing_sheets:
            lines.append(f"不足シート: {', '.join(ss.missing_sheets)}")
        if ss.unexpected_sheets:
            lines.append(f"余分なシート: {', '.join(ss.unexpected_sheets)}")
        for sheet, counts in ss.row_counts.items():
            exp, act = counts.get("expected"), counts.get("actual")
            mark = "OK" if exp == act else f"NG(期待:{exp} 実際:{act})"
            lines.append(f"  シート '{sheet}' 行数: {mark}")
        for sheet, vis in ss.visibility.items():
            exp, act = vis.get("expected"), vis.get("actual")
            mark = "OK" if exp == act else f"NG(期待:{exp} 実際:{act})"
            lines.append(f"  シート '{sheet}' 表示状態: {mark}")

        lines.append("\n[キーセル比較]")
        for r in self.key_cell_results:
            mark = "OK" if r.match else f"NG(期待:{r.expected!r} 実際:{r.actual!r})"
            lines.append(f"  {r.coord} {r.description}: {mark}")

        lines.append("\n[値ベーススキャン]")
        for r in self.value_scan_results:
            if not r.found:
                lines.append(f"  '{r.label}' ({r.description}): ラベル未発見")
                continue
            lines.append(f"  '{r.label}' ({r.description}): 発見 at {r.found_at}")
            for f in r.field_results:
                mark = "OK" if f.match else f"NG(期待:{f.expected} 実際:{f.actual})"
                lines.append(f"    {f.name}: {mark}")

        lines.append("\n[色チェック]")
        for r in self.color_check_results:
            mark = "OK" if r.passed else f"NG({r.cells_matched}/{r.cells_checked}セル一致)"
            lines.append(f"  {r.description}: {mark}")

        lines.append("\n[追加ファイル]")
        for r in self.extra_file_results:
            mark = "OK" if r.get("found") else "NG(ファイル未作成)"
            lines.append(f"  {r.get('description', r.get('pattern'))}: {mark}")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _values_close(a: object, b: object, tolerance: float = 0.05) -> bool:
    """Return True if a and b are numerically close or string-equal."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        fa, fb = float(a), float(b)
        if fb == 0:
            return abs(fa) < 1e-9
        return abs(fa - fb) / abs(fb) <= tolerance
    except (TypeError, ValueError):
        return str(a).strip() == str(b).strip()


def _parse_rgb(rgb_str: str) -> tuple[int, int, int] | None:
    """Parse an 8-char ARGB or 6-char RGB hex string into (r, g, b)."""
    s = rgb_str.lstrip("#").upper()
    if len(s) == 8:   # AARRGGBB
        s = s[2:]
    if len(s) == 6:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
        return r, g, b
    return None


def _cell_color_family(cell) -> str | None:
    """Return the color family name of a cell's fill, or None if uncolored."""
    try:
        fill = cell.fill
        if fill is None or fill.fill_type in (None, "none"):
            return None
        rgb_str = fill.fgColor.rgb
        if not rgb_str or rgb_str in ("00000000", "FFFFFFFF", "00FFFFFF"):
            return None
        parsed = _parse_rgb(rgb_str)
        if parsed is None:
            return None
        r, g, b = parsed
        for name, predicate in _COLOR_FAMILIES.items():
            if predicate(r, g, b):
                return name
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Comparison functions
# ---------------------------------------------------------------------------


def _compare_sheets(
    wb_actual: openpyxl.Workbook,
    wb_expected: openpyxl.Workbook,
    visibility_checks: dict[str, str] | None = None,
) -> SheetStructureResult:
    expected_names = set(wb_expected.sheetnames)
    actual_names = set(wb_actual.sheetnames)

    row_counts: dict[str, dict] = {}
    for name in expected_names:
        exp_rows = wb_expected[name].max_row if name in wb_expected.sheetnames else None
        act_rows = wb_actual[name].max_row if name in actual_names else None
        row_counts[name] = {"expected": exp_rows, "actual": act_rows}

    visibility: dict[str, dict] = {}
    if visibility_checks:
        for sheet_name, expected_state in visibility_checks.items():
            if sheet_name in wb_actual.sheetnames:
                actual_state = wb_actual[sheet_name].sheet_state
            else:
                actual_state = "missing"
            visibility[sheet_name] = {"expected": expected_state, "actual": actual_state}

    return SheetStructureResult(
        missing_sheets=sorted(expected_names - actual_names),
        unexpected_sheets=sorted(actual_names - expected_names),
        row_counts=row_counts,
        visibility=visibility,
    )


def _compare_key_cells(
    wb_actual: openpyxl.Workbook,
    wb_expected: openpyxl.Workbook,
    key_cells: dict[str, list[dict]],
) -> list[KeyCellResult]:
    results: list[KeyCellResult] = []
    for sheet_name, cells in key_cells.items():
        if sheet_name not in wb_actual.sheetnames or sheet_name not in wb_expected.sheetnames:
            for cell_def in cells:
                results.append(KeyCellResult(
                    coord=cell_def["coord"],
                    description=cell_def.get("description", ""),
                    expected=None,
                    actual=None,
                    match=False,
                ))
            continue

        ws_a = wb_actual[sheet_name]
        ws_e = wb_expected[sheet_name]
        for cell_def in cells:
            coord = cell_def["coord"]
            tolerance = cell_def.get("tolerance", 0.05)
            expected = ws_e[coord].value
            actual = ws_a[coord].value
            results.append(KeyCellResult(
                coord=coord,
                description=cell_def.get("description", ""),
                expected=expected,
                actual=actual,
                match=_values_close(actual, expected, tolerance),
            ))
    return results


def _value_scan(
    wb_actual: openpyxl.Workbook,
    scans: dict[str, list[dict]],
) -> list[ValueScanResult]:
    """Scan sheets by label value, not by fixed row position."""
    results: list[ValueScanResult] = []

    for sheet_name, scan_list in scans.items():
        if sheet_name not in wb_actual.sheetnames:
            for scan in scan_list:
                results.append(ValueScanResult(
                    label=scan["label"],
                    description=scan.get("description", scan["label"]),
                    found=False, found_at=None, field_results=[],
                ))
            continue

        ws = wb_actual[sheet_name]

        for scan in scan_list:
            label = scan["label"]
            desc = scan.get("description", label)
            adj_fields = scan.get("adjacent_fields", [])

            # Search entire sheet for a cell matching the label
            found_row: int | None = None
            found_col: int | None = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() == label.strip():
                        found_row = cell.row
                        found_col = cell.column
                        break
                if found_row is not None:
                    break

            if found_row is None:
                results.append(ValueScanResult(
                    label=label, description=desc,
                    found=False, found_at=None, field_results=[],
                ))
                continue

            from openpyxl.utils import get_column_letter
            found_at = f"{get_column_letter(found_col)}{found_row}"
            field_results: list[ValueScanFieldResult] = []

            for adj in adj_fields:
                offset = adj["offset_col"]
                name = adj["name"]
                expected = adj.get("expected")
                tolerance = adj.get("tolerance", 0.05)
                actual = ws.cell(found_row, found_col + offset).value
                field_results.append(ValueScanFieldResult(
                    name=name,
                    expected=expected,
                    actual=actual,
                    match=_values_close(actual, expected, tolerance),
                ))

            results.append(ValueScanResult(
                label=label, description=desc,
                found=True, found_at=found_at,
                field_results=field_results,
            ))

    return results


def _check_colors(
    wb_actual: openpyxl.Workbook,
    color_checks: dict[str, list[dict]],
) -> list[ColorCheckResult]:
    results: list[ColorCheckResult] = []

    for sheet_name, checks in color_checks.items():
        if sheet_name not in wb_actual.sheetnames:
            for check in checks:
                r1, r2 = check["rows"]
                c1, c2 = check["cols"]
                results.append(ColorCheckResult(
                    description=check.get("description", ""),
                    row_range=(r1, r2), col_range=(c1, c2),
                    expected_color_family=check.get("color_family", "any"),
                    cells_checked=0, cells_matched=0,
                    match_rate=0.0, passed=False,
                ))
            continue

        ws = wb_actual[sheet_name]
        for check in checks:
            r1, r2 = check["rows"]
            c1, c2 = check["cols"]
            color_family = check.get("color_family", "any")
            threshold = check.get("threshold", 0.5)

            cells_checked = 0
            cells_matched = 0
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    cell = ws.cell(r, c)
                    cells_checked += 1
                    family = _cell_color_family(cell)
                    if color_family == "any":
                        if family is not None:
                            cells_matched += 1
                    elif family == color_family:
                        cells_matched += 1

            match_rate = cells_matched / cells_checked if cells_checked else 0.0
            results.append(ColorCheckResult(
                description=check.get("description", ""),
                row_range=(r1, r2), col_range=(c1, c2),
                expected_color_family=color_family,
                cells_checked=cells_checked,
                cells_matched=cells_matched,
                match_rate=round(match_rate, 3),
                passed=match_rate >= threshold,
            ))

    return results


def _check_extra_files(
    extra_files: list[dict],
    search_dir: str = ".",
) -> list[dict]:
    results: list[dict] = []
    for spec in extra_files:
        pattern = spec["pattern"]
        desc = spec.get("description", pattern)
        matches = glob.glob(str(Path(search_dir) / pattern))
        results.append({
            "pattern": pattern,
            "description": desc,
            "found": len(matches) > 0,
            "files": matches,
        })
    return results


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def compare_excel_structured(
    actual_path: str,
    expected_path: str,
    rubric: dict | None = None,
    extra_file_search_dir: str = ".",
) -> StructuredCompareReport:
    """Compare actual vs expected Excel using structured rubric.

    Args:
        actual_path: Path to the generated output file.
        expected_path: Path to the reference/correct file.
        rubric: Task-specific evaluation rubric dict (from rubric.json).
                If None, only sheet structure is compared.
        extra_file_search_dir: Directory to search for extra output files.

    Returns:
        StructuredCompareReport with all comparison results.
    """
    rubric = rubric or {}

    try:
        wb_actual = openpyxl.load_workbook(actual_path, data_only=True)
        wb_expected = openpyxl.load_workbook(expected_path, data_only=True)
    except Exception as exc:
        logger.warning("Failed to load workbooks: %s", exc)
        empty = SheetStructureResult(
            missing_sheets=[], unexpected_sheets=[],
            row_counts={}, visibility={},
        )
        return StructuredCompareReport(
            sheet_structure=empty,
            key_cell_results=[], value_scan_results=[],
            color_check_results=[], extra_file_results=[],
        )

    visibility_checks: dict[str, str] = {}
    for sheet in rubric.get("sheet_visibility", {}).get("hidden", []):
        visibility_checks[sheet] = "hidden"
    for sheet in rubric.get("sheet_visibility", {}).get("visible", []):
        visibility_checks[sheet] = "visible"

    sheet_structure = _compare_sheets(wb_actual, wb_expected, visibility_checks)
    key_cell_results = _compare_key_cells(wb_actual, wb_expected, rubric.get("key_cells", {}))
    value_scan_results = _value_scan(wb_actual, rubric.get("value_scan", {}))
    color_check_results = _check_colors(wb_actual, rubric.get("color_checks", {}))
    extra_file_results = _check_extra_files(
        rubric.get("extra_files", []), search_dir=extra_file_search_dir
    )

    return StructuredCompareReport(
        sheet_structure=sheet_structure,
        key_cell_results=key_cell_results,
        value_scan_results=value_scan_results,
        color_check_results=color_check_results,
        extra_file_results=extra_file_results,
    )
