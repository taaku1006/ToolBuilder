"""Stage 1: UNDERSTAND — Analyze file structure and plan strategy.

ExcelAnalyzer performs LLM-free analysis using openpyxl/pandas directly.
StrategyPhase makes a single LLM call (gpt-4o-mini) to classify task
complexity and decide the generation approach.
"""

from __future__ import annotations

import json
import logging
import os
from dataclasses import asdict
from pathlib import Path

import openpyxl

from excel.xlsx_parser import SheetInfo, parse_file, build_file_context
from infra.openai_client import OpenAIClient
from infra.prompt_loader import load_prompt
from pipeline.v2.config import STAGE_CONFIGS
from pipeline.v2.models import (
    ComplexitySignals,
    FileContext,
    MemoryContext,
    Strategy,
    StrategyStep,
    TaskClassification,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# ExcelAnalyzer — LLM-free file analysis
# ---------------------------------------------------------------------------


class ExcelAnalyzer:
    """Analyze Excel file structure without any LLM calls.

    Reuses xlsx_parser.parse_file() for basic info, then adds
    deeper structural signals (formulas, charts, pivot tables,
    complexity indicators).
    """

    def analyze(self, file_path: str) -> FileContext:
        sheets_raw = parse_file(file_path)
        sheets_dicts = [_sheet_info_to_dict(s) for s in sheets_raw]

        has_merged = any(s.merged_cells for s in sheets_raw)
        has_formulas = False
        has_charts = False
        has_pivot = False

        ext = Path(file_path).suffix.lower()
        if ext in (".xlsx", ".xlsm"):
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False)
                has_formulas = self._detect_formulas(wb)
                has_charts = self._detect_charts(wb)
                has_pivot = self._detect_pivot_tables(wb)
                wb.close()
            except Exception:
                logger.warning("Failed to analyze workbook details", exc_info=True)

        complexity = self._compute_complexity(sheets_raw, has_merged, has_formulas)

        return FileContext(
            sheets=sheets_dicts,
            has_merged_cells=has_merged,
            has_formulas=has_formulas,
            has_charts=has_charts,
            has_pivot_tables=has_pivot,
            file_size_mb=os.path.getsize(file_path) / (1024 * 1024),
            complexity_signals=complexity,
        )

    @staticmethod
    def _detect_formulas(wb: openpyxl.Workbook) -> bool:
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=min(ws.max_row or 0, 50)):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        return True
        return False

    @staticmethod
    def _detect_charts(wb: openpyxl.Workbook) -> bool:
        for ws in wb.worksheets:
            if ws._charts:  # noqa: SLF001
                return True
        return False

    @staticmethod
    def _detect_pivot_tables(wb: openpyxl.Workbook) -> bool:
        for ws in wb.worksheets:
            if hasattr(ws, "pivotTables") and ws.pivotTables:
                return True
        return False

    @staticmethod
    def _compute_complexity(
        sheets: list[SheetInfo],
        has_merged: bool,
        has_formulas: bool,
    ) -> ComplexitySignals:
        multi_sheet = len(sheets) > 1

        # Detect mixed dtypes per column
        mixed_cols: list[str] = []
        for s in sheets:
            for col, dtype in (s.types or {}).items():
                if dtype == "mixed" or "/" in str(dtype):
                    mixed_cols.append(f"{s.name}.{col}")

        # Detect nested headers (heuristic: first row has merged cells)
        nested = has_merged  # simplified heuristic

        total_cells = sum(
            (s.total_rows or 0) * len(s.headers or []) for s in sheets
        )

        return ComplexitySignals(
            multi_sheet_refs=multi_sheet,
            nested_headers=nested,
            mixed_dtypes_per_column=mixed_cols,
            estimated_total_cells=total_cells,
        )


def _sheet_info_to_dict(s: SheetInfo) -> dict:
    """Convert a SheetInfo dataclass to a plain dict."""
    return {
        "name": s.name,
        "total_rows": s.total_rows,
        "headers": list(s.headers) if s.headers else [],
        "types": dict(s.types) if s.types else {},
        "preview": list(s.preview) if s.preview else [],
        "merged_cells": tuple(s.merged_cells) if s.merged_cells else (),
    }


# ---------------------------------------------------------------------------
# StrategyPhase — single LLM call for task classification + strategy
# ---------------------------------------------------------------------------


class StrategyPhase:
    """Decide the code generation strategy before writing any code."""

    def __init__(self, openai_client: OpenAIClient, settings=None) -> None:
        self._client = openai_client
        self._settings = settings

    async def plan(
        self,
        task: str,
        file_context: FileContext,
        memory_context: MemoryContext,
    ) -> tuple[TaskClassification, Strategy]:
        prompt_template = load_prompt("v2_strategize", self._settings)
        prompt = prompt_template.format(
            task=task,
            file_context=file_context.to_prompt(),
            past_patterns=memory_context.to_prompt() if memory_context.patterns else "なし",
            past_gotchas=memory_context.to_prompt() if memory_context.gotchas else "なし",
        )

        cfg = STAGE_CONFIGS["strategize"]
        raw = self._client.chat(
            [{"role": "user", "content": prompt}],
            model=cfg["model"],
            temperature=cfg["temperature"],
            max_tokens=cfg["max_tokens"],
        )

        return _parse_strategy_response(raw)

    async def replan(
        self,
        task: str,
        file_context: FileContext,
        memory_context: MemoryContext,
        previous_strategy: Strategy,
        failure_info: list,
    ) -> tuple[TaskClassification, Strategy]:
        """Re-plan with information about what failed."""
        failure_summary = "\n".join(
            f"- {a.approach}: {a.error_message or 'quality insufficient'}"
            for a in failure_info[-3:]
        )
        extra_context = (
            f"\n## 前回の失敗\n前回のアプローチ ({previous_strategy.approach}) は"
            f"失敗しました。\n{failure_summary}\n異なるアプローチを提案してください。"
        )

        prompt_template = load_prompt("v2_strategize", self._settings)
        prompt = prompt_template.format(
            task=task + extra_context,
            file_context=file_context.to_prompt(),
            past_patterns=memory_context.to_prompt() if memory_context.patterns else "なし",
            past_gotchas=memory_context.to_prompt() if memory_context.gotchas else "なし",
        )

        cfg = STAGE_CONFIGS["strategize"]
        raw = self._client.chat(
            [{"role": "user", "content": prompt}],
            model=cfg["model"],
            temperature=cfg["temperature"],
            max_tokens=cfg["max_tokens"],
        )

        return _parse_strategy_response(raw)


def _extract_json(raw: str) -> str:
    """Extract JSON from LLM response that may contain surrounding text."""
    text = raw.strip()

    # Try 1: Strip markdown code fences
    import re
    fence_match = re.search(r"```(?:json)?\s*\n?(.*?)\n?\s*```", text, re.DOTALL)
    if fence_match:
        return fence_match.group(1).strip()

    # Try 2: Find first { ... } block
    first_brace = text.find("{")
    if first_brace >= 0:
        # Find matching closing brace
        depth = 0
        for i in range(first_brace, len(text)):
            if text[i] == "{":
                depth += 1
            elif text[i] == "}":
                depth -= 1
                if depth == 0:
                    return text[first_brace:i + 1]

    return text


def _parse_strategy_response(raw: str) -> tuple[TaskClassification, Strategy]:
    """Parse LLM JSON response into TaskClassification + Strategy."""
    text = _extract_json(raw)

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        logger.warning(
            "Failed to parse strategy response, using defaults. Raw (first 200): %s",
            raw[:200],
        )
        return TaskClassification(), Strategy()

    classification = TaskClassification(
        complexity=data.get("complexity", "standard"),
        task_type=data.get("task_type", "general"),
        estimated_difficulty=data.get("estimated_difficulty", 0.5),
    )

    steps = None
    if data.get("steps"):
        steps = [
            StrategyStep(
                id=s.get("id", i + 1),
                action=s.get("action", ""),
                verify=s.get("verify", ""),
                expected_in_output=s.get("expected_in_output", []),
            )
            for i, s in enumerate(data["steps"])
        ]

    strategy = Strategy(
        approach=data.get("library", "pandas"),
        key_functions=data.get("key_functions", []),
        preprocessing_steps=data.get("preprocessing", []),
        output_format=data.get("output_format", "xlsx"),
        risk_factors=data.get("risk_factors", []),
        steps=steps,
    )

    return classification, strategy
